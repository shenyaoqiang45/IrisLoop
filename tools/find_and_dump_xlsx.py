"""Locate the protocol xlsx in Temp via globs and dump it to text.

Avoids PowerShell Chinese-path encoding issues: all path handling stays in Python.
"""

from __future__ import annotations

import glob
import os
import sys

import openpyxl


def cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\r\n", "\n").replace("\n", " ⏎ ")
    return str(v)


def dump(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src, data_only=True)
    lines: list[str] = []
    lines.append(f"FILE: {src}")
    lines.append(f"SHEETS({len(wb.sheetnames)}): {wb.sheetnames}")

    for ws in wb.worksheets:
        lines.append("")
        lines.append("=" * 78)
        lines.append(f"### SHEET: {ws.title}  max_row={ws.max_row} max_col={ws.max_column}")
        lines.append("=" * 78)
        for row in ws.iter_rows():
            vals = [cell_text(c.value).strip() for c in row]
            if not any(vals):
                continue
            lines.append(f"R{row[0].row:>4}| " + " | ".join(vals))

    with open(dst, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print("SHEETS:", wb.sheetnames)
    print("WRITTEN:", dst)


def main() -> int:
    dst = sys.argv[1] if len(sys.argv) > 1 else r"f:\2026\IrisLoop\tools\ble_protocol_dump.txt"

    # Search Temp first, then common download folders
    temp = os.environ.get("TEMP", r"C:\Windows\Temp")
    roots = [
        os.path.join(temp, "codebuddy-dropped-files"),
        temp,
        os.path.join(os.path.expanduser("~"), "Downloads"),
    ]

    found: list[str] = []
    for root in roots:
        if not os.path.isdir(root):
            continue
        # Keep Chinese globs so vendor filenames still match; also search English *Iris*.xlsx
        for pat in ("*蓝牙*.xlsx", "*协议*.xlsx", "*Iris*.xlsx"):
            found.extend(glob.glob(os.path.join(root, "**", pat), recursive=True))

    # Dedupe
    seen = set()
    uniq = []
    for p in found:
        rp = os.path.realpath(p)
        if rp not in seen:
            seen.add(rp)
            uniq.append(p)

    print("CANDIDATES:")
    for p in uniq:
        print("  ", p, os.path.getsize(p))

    if not uniq:
        print("[error] protocol xlsx not found")
        return 1

    # Prefer names containing 协议/蓝牙 (vendor Chinese filenames)
    target = None
    for p in uniq:
        b = os.path.basename(p)
        if "协议" in b or "蓝牙" in b:
            target = p
            break
    target = target or uniq[0]

    print("\nUSING:", target)
    dump(target, dst)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
