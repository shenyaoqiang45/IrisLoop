"""Dump every sheet of an xlsx to plain text for full reading.

Usage: python tools/xlsx_dump.py <xlsx-path> [output-txt-path]
"""

from __future__ import annotations

import sys

import openpyxl


def cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\r\n", "\n").replace("\n", " ⏎ ")
    return str(v)


def main() -> int:
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else src + ".txt"

    wb = openpyxl.load_workbook(src, data_only=True)
    lines: list[str] = []
    lines.append(f"FILE: {src}")
    lines.append(f"SHEETS: {len(wb.sheetnames)}")

    for ws in wb.worksheets:
        lines.append("")
        lines.append("=" * 78)
        lines.append(f"### SHEET: {ws.title}   dims={ws.dimensions} "
                     f"max_row={ws.max_row} max_col={ws.max_column}")
        lines.append("=" * 78)

        for row in ws.iter_rows():
            vals = [cell_text(c.value).strip() for c in row]
            if not any(vals):
                continue
            lines.append(f"R{row[0].row:>4}| " + " | ".join(vals))

    with open(dst, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"written: {dst}")
    print(f"sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
