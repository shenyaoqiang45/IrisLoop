"""Convert doc/Iris Green蓝牙通信协议_20260324_V1.9.xlsx to Markdown under doc/.

Usage: python tools/xlsx_to_md.py
"""

from __future__ import annotations

import os

import openpyxl

SRC = r"doc\Iris Green蓝牙通信协议_20260324_V1.9.xlsx"
DST = r"doc\Iris Green蓝牙通信协议_20260324_V1.9.md"


def cell_str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    # Newlines inside cells become <br>; escape pipes so markdown tables stay intact
    return s.replace("|", "\\|").replace("\n", "<br>").strip()


def sheet_to_md(ws, out: list[str]) -> None:
    # Fill merged cells with the top-left value only
    merged: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = v

    def get(r: int, c: int) -> str:
        v = ws.cell(row=r, column=c).value
        if v is None and (r, c) in merged:
            v = merged[(r, c)]
        return cell_str(v)

    def dedup(row: list[str]) -> list[str]:
        """Horizontal merge fill repeats values; keep only the first occurrence in a row."""
        seen: set[str] = set()
        out: list[str] = []
        for v in row:
            if v and v in seen:
                out.append("")
            else:
                if v:
                    seen.add(v)
                out.append(v)
        return out

    max_r, max_c = ws.max_row, ws.max_column
    # Trim all-empty trailing columns
    while max_c > 1 and all(not get(r, max_c) for r in range(1, max_r + 1)):
        max_c -= 1

    for r in range(1, max_r + 1):
        row = dedup([get(r, c) for c in range(1, max_c + 1)])
        # Strip consecutive empty columns at both ends
        while row and not row[0]:
            row.pop(0)
        while row and not row[-1]:
            row.pop()
        if not row:
            continue
        # One non-empty cell -> paragraph; several cells -> table row
        nonempty = [i for i, v in enumerate(row) if v]
        if len(nonempty) <= 1:
            v = row[nonempty[0]] if nonempty else ""
            if v.isdigit():  # leftover numeric row index; drop
                continue
            out.append(v)
            out.append("")
        else:
            out.append("| " + " | ".join(row) + " |")
            # After the first row of a contiguous table block, insert the separator
            if len(out) < 2 or not out[-2].startswith("|"):
                ncols = len(row)
                out.insert(len(out) - 1, "|" + "---|" * ncols)


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out: list[str] = [
        "# Iris Green Bluetooth Protocol V1.9 (2026-03-24)",
        "",
        f"> Auto-exported from `{os.path.basename(SRC)}` by tools/xlsx_to_md.py; for reading only.",
        "> Edit the xlsx source and re-export when the protocol changes.",
        "",
    ]
    for name in wb.sheetnames:
        ws = wb[name]
        out.append(f"## Sheet: {name}")
        out.append("")
        sheet_to_md(ws, out)
        out.append("")

    text = "\n".join(out)
    # Collapse 3+ consecutive blank lines
    while "\n\n\n" in text:
        text = text.replace("\n\n\n", "\n\n")
    with open(DST, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"written {DST}  {len(text)} chars")


if __name__ == "__main__":
    main()
